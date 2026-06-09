# utils/excel_report.py

import os
import shutil
import openpyxl
from datetime import datetime


def generate_excel_report(asset: dict, results: list) -> str:
    """
    asset: 자산 정보 딕셔너리
    results: 점검 결과 리스트
    반환: 생성된 Excel 파일 경로
    """
    template_path = "template/리눅스_template.xlsx"
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    # 출력 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"{output_dir}/{asset['code']}_{timestamp}.xlsx"

    # 템플릿 복사 (원본 보존)
    shutil.copy(template_path, output_path)

    # 복사본 열기
    wb = openpyxl.load_workbook(output_path)

    # 각 시트 데이터 채우기
    _fill_asset_sheet(wb, asset)
    _fill_stats_sheet(wb, results)
    _fill_detail_sheet(wb, asset, results)

    wb.save(output_path)
    return output_path


def _fill_asset_sheet(wb, asset: dict):
    """자산목록 시트에 자산 정보 입력"""
    try:
        ws = wb["자산목록"]
    except KeyError:
        return

    ws.cell(row=2, column=1, value=asset.get("code", ""))
    ws.cell(row=2, column=2, value=asset.get("hostname", ""))
    ws.cell(row=2, column=3, value=asset.get("ip", ""))
    ws.cell(row=2, column=4, value=asset.get("os", ""))
    ws.cell(row=2, column=5, value=asset.get("version", ""))
    ws.cell(row=2, column=6, value=asset.get("purpose", ""))
    ws.cell(row=2, column=7, value=asset.get("location", ""))
    ws.cell(row=2, column=8, value=asset.get("manager", ""))


def _fill_stats_sheet(wb, results: list):
    """항목별 통계 시트에 결과값 입력"""
    try:
        ws = wb["항목별 통계"]
    except KeyError:
        return

    result_map = {r["checklist_code"]: r["result"] for r in results}

    for row in ws.iter_rows(min_row=2):
        code_cell = row[0].value
        if code_cell and code_cell in result_map:
            row[1].value = result_map[code_cell]


def _fill_detail_sheet(wb, asset: dict, results: list):
    """장비별 상세 시트에 점검 결과 입력"""
    sheet_name = asset.get("code", "").replace("-", "_")

    try:
        ws = wb[sheet_name]
    except KeyError:
        return

    result_map = {r["checklist_code"]: r for r in results}

    for row in ws.iter_rows(min_row=2):
        code_cell = row[0].value
        if code_cell and code_cell in result_map:
            r = result_map[code_cell]
            row[1].value = r.get("result", "")
            row[2].value = r.get("current_status", "")
            row[3].value = r.get("improvement", "")
